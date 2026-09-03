"""Read the NUML workbook and turn it into clean JSONL files."""

import json
import re
from pathlib import Path
from openpyxl import load_workbook

# Adapted from the chapter's absolute macOS path to this project's own
# data/raw folder (see Chapter 2 project structure).
WORKBOOK = "data/raw/NUML_data_collection_template.xlsx"

KNOWLEDGE_REQUIRED = ["id", "category", "text", "status"]
KNOWLEDGE_OPTIONAL = ["title", "source", "written_by"]

INSTRUCTION_REQUIRED = ["id", "category", "instruction", "output", "status"]
INSTRUCTION_OPTIONAL = ["source_ids", "written_by"]

STATUS_ALIASES = {
    "verified": "verified", "approved": "verified", "ok": "verified",
    "checked": "verified", "done": "verified",
    "draft": "draft", "pending": "draft", "": "draft",
    "rejected": "rejected", "reject": "rejected",
}


def normalise_status(raw: str) -> str:
    """Map whatever a student typed onto our three official values."""
    return STATUS_ALIASES.get((raw or "").strip().lower(), "unknown")


def normalise_header(raw) -> str:
    """'text (100-300 words, facts only)' becomes 'text'."""
    if raw is None:
        return ""
    return str(raw).split("(")[0].strip().lower().replace(" ", "_")


def clean(text) -> str:
    """Remove characters that Word and Google Docs sneak in."""
    if text is None:
        return ""
    text = str(text)
    swaps = {"\u2018": "'", "\u2019": "'", "\u201c": '"', "\u201d": '"',
             "\u2013": "-", "\u2014": "-", "\u00a0": " ", "\ufeff": "", "\u2022": "-"}
    for bad, good in swaps.items():
        text = text.replace(bad, good)
    return re.sub(r"\s+", " ", text).strip()


def make_title(text: str, max_words: int = 9) -> str:
    """Build a title from the first sentence when none was written."""
    if not text:
        return "Untitled"
    first = re.split(r"(?<=[.!?])\s", text.strip())[0]
    words = first.split()
    return " ".join(words[:max_words]) + ("..." if len(words) > max_words else "")


def read_tab(workbook_path: str, tab: str, required: list[str], optional: list[str]) -> list[dict]:
    """EXTRACT. Crash on missing required columns, warn on missing optional ones."""
    wb = load_workbook(workbook_path, data_only=True)

    if tab not in wb.sheetnames:
        raise ValueError(f"Tab '{tab}' not found. Tabs present: {wb.sheetnames}")
    ws = wb[tab]

    headers = [normalise_header(c.value) for c in ws[1]]

    missing = [f for f in required if f not in headers]
    if missing:
        raise ValueError(f"Tab '{tab}' is missing REQUIRED columns: {missing}. Found: {headers}")

    absent = [f for f in optional if f not in headers]
    if absent:
        print(f"  note: optional columns not in '{tab}': {absent}")

    rows = []
    for excel_row in range(2, ws.max_row + 1):
        values = {name: clean(ws.cell(row=excel_row, column=col).value)
                  for col, name in enumerate(headers, start=1) if name}
        if not any(values.values()):
            continue
        for field in optional:
            values.setdefault(field, "")
        values["_row"] = excel_row
        rows.append(values)
    return rows


def to_knowledge(rows: list[dict]) -> tuple[list[dict], list[tuple]]:
    """TRANSFORM. Keep verified rows, report the rest."""
    kept, skipped = [], []
    for r in rows:
        status = normalise_status(r.get("status"))
        if status != "verified":
            skipped.append((status, r["_row"], r.get("id") or "no id", r.get("status") or "blank"))
            continue
        kept.append({
            "id": r["id"],
            "category": r["category"],
            "title": r["title"] or make_title(r["text"]),
            "text": r["text"],
            "source": r["source"] or "not recorded",
            "written_by": r["written_by"] or "unknown",
            "origin": "handwritten",
        })
    return kept, skipped


def to_instructions(rows: list[dict]) -> tuple[list[dict], list[tuple]]:
    kept, skipped = [], []
    for r in rows:
        status = normalise_status(r.get("status"))
        if status != "verified":
            skipped.append((status, r["_row"], r.get("id") or "no id", r.get("status") or "blank"))
            continue
        kept.append({
            "id": r["id"],
            "category": r["category"],
            "instruction": r["instruction"],
            "output": r["output"],
            "source_ids": [s.strip() for s in r["source_ids"].split(",") if s.strip()],
            "written_by": r["written_by"] or "unknown",
        })
    return kept, skipped


def save_jsonl(rows: list[dict], path: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def load_jsonl(path: str) -> list[dict]:
    p = Path(path)
    if not p.exists():
        return []
    return [json.loads(l) for l in p.read_text(encoding="utf-8").splitlines() if l.strip()]